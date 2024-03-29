import telebot
from openpyxl import load_workbook
import requests
from io import BytesIO
from telebot import types

# Инициализация бота
bot = telebot.TeleBot('7123896415:AAHaVQeFj4Pc9_zlxi3DuprdRH9RIO1qoK4')

# ID файла Excel на Google Диске (из его общедоступной ссылки)
file_id = '1fo3Wbb6noDILXL96qkodV4uVInZYZY-p'

# SKU и тип цены, введенные пользователем
#user_price_type = None
user_sku = None

# Функция загрузки данных из Excel-файла, расположенного на Google Диске
def load_excel_data_from_google_drive(file_id):
    url = f'https://drive.google.com/uc?export=download&id={file_id}'
    response = requests.get(url)
    workbook = load_workbook(filename=BytesIO(response.content))
    sheet = workbook.active
    data = {}
    for row in sheet.iter_rows(values_only=True):
        sku = row[0]
        if sku == user_sku:  # Проверяем, соответствует ли SKU введенному пользователем
            price_column = 2 if user_price_type == sheet.cell(row=1, column=2).value else 3
            price = row[price_column - 1]
            comment = row[3] if len(row) >= 4 else None
            name = row[4] if len(row) >= 5 else None  # Добавляем данные из четвертого столбца "Name"
            data[sku] = {'price': price, 'comment': comment, 'name': name}  # Сохраняем все данные в словаре
            break
    return data

# Обработка команды /start
@bot.message_handler(commands=['start'])
def send_welcome(message):
    global user_price_type, user_sku
    user_price_type = None
    user_sku = None
    reply = "Привет! Я бот, который поможет тебе найти значения цен и комментариев по Коду Услуги и Региону.\n" \
            "Выберите нужный вам регион: Moscow, Region"
    #bot.reply_to(message, reply)
    bot.send_message(message.chat.id, text=reply, reply_markup=create_price_type_keyboard())

# Функция создания клавиатуры с типами цен
def create_price_type_keyboard():
    keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    price_types = ['Moscow', 'Region']  # Заголовки таблицы
    keyboard.add(*[types.KeyboardButton(price_type) for price_type in price_types])
    return keyboard

# Обработка текстовых сообщений
@bot.message_handler(func=lambda message: True)
def handle_message(message):
    global user_price_type, user_sku
    if user_price_type is None:
        if message.text not in ['Moscow', 'Region']:  # Проверяем, что тип цены входит в список
            bot.reply_to(message, "Выберите нужный вам регион: Moscow, Region")
            return
        user_price_type = message.text
        bot.reply_to(message, "Теперь введите Код Услуги ДЛ (пожалуйста, сохраняйте регистр букв):")
    elif user_sku is None:
        user_sku = message.text
        excel_data = load_excel_data_from_google_drive(file_id)
        if user_sku in excel_data:
            price = excel_data[user_sku]['price']
            comment = excel_data[user_sku]['comment']
            name = excel_data[user_sku]['name']  # Получаем данные из четвертого столбца "Name"
            if comment:
                bot.reply_to(message, f"Базовая цена для услуги {user_sku} {name} (регион: {user_price_type}): {price} руб. \nКомментарий: {comment}. \nЗапросить цену на другую услугу /start")
            else:
                bot.reply_to(message, f"Базовая цена для услуги {user_sku} {name} (регион: {user_price_type}): {price} руб. \nЗапросить цену на другую услугу /start")
        else:
            bot.reply_to(message, f"Код услуги {user_sku} не найден. \nПовторите запрос, нажав /start")


# Запуск бота
from urllib3.exceptions import ProtocolError

try:
    bot.polling()
except ProtocolError as pe:
    print(f"A ProtocolError occurred: {pe}")
except Exception as e:
    print(f"An error occurred: {e}")
